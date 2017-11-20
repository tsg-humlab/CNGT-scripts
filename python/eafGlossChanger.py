#!/usr/bin/python

"""
Script to add Signbank as a lexicon to the CNGT EAFs.
"""

import sys
import getopt
from openpyxl import load_workbook
from filecollectionprocessing.filecollectionprocessor import FileCollectionProcessor
from filecollectionprocessing.eafprocessor import EafProcessor

# Settings


class GlossChanger(EafProcessor):
    """
    
    """

    def __init__(self, excel_with_changes, first_row = 1, old_gloss_column = 'A', new_gloss_column = 'B',
                 meaning_column = 'C'):
        # Reading the Excel with changes
        self.changes = dict()
        self.read_excel_with_changes(excel_with_changes, first_row, old_gloss_column, new_gloss_column, meaning_column,
                                     self.changes)

    @staticmethod
    def read_excel_with_changes(excel_with_changes, first_row, old_gloss_column, new_gloss_column,
                                meaning_column, changes):
        """
        Reads the Excel containing changes and puts them in a dictionary
        :param excel_with_changes: 
        :param first_row: 
        :param old_gloss_column: 
        :param new_gloss_column: 
        :param meaning_column: 
        :return: 
        """
        # Load the workbook
        workbook = load_workbook(filename=excel_with_changes)
        # Take the first worksheet
        worksheet = workbook[workbook.get_sheet_names()[0]]
        for row_index in range(first_row, worksheet.max_row+1):
            try:
                old_gloss = worksheet[old_gloss_column+str(row_index)].value
                new_gloss = worksheet[new_gloss_column+str(row_index)].value
                meaning = worksheet[meaning_column+str(row_index)].value
                if (old_gloss is not None and old_gloss is not "" and
                    new_gloss is not None and new_gloss is not "" and
                    meaning is not None and meaning is not "" and
                    old_gloss not in changes):
                    changes[old_gloss] = (new_gloss, meaning)
            except Exception:
                print("Unable to process row " + str(row_index) + ". Error message: " + sys.exc_info())

    def process_eaf(self, eaf, file_name):
        print("EAF file: " + file_name)
        for subject_id in [1, 2]:
            for hand in ['L', 'R']:
                # Handle Meaning tier
                meaning_tier_id = 'Meaning' + hand + ' S' + str(subject_id)
                meaning_tier = eaf.tiers[meaning_tier_id]
                meaning_annotation_dict = {}
                for meaning_ann_id, meaning_ann_contents in meaning_tier[1].items():
                    # Switch annotation id and refering id. This is possible because
                    # the stereo type of Meaning tiers is symbolic association.
                    meaning_annotation_dict[meaning_ann_contents[0]] = (meaning_ann_id,
                                                                       meaning_ann_contents[1],
                                                                       meaning_ann_contents[2],
                                                                       meaning_ann_contents[3],
                                                                       )

                # Handle Gloss tier
                gloss_tier_id = 'Gloss' + hand + ' S' + str(subject_id)
                gloss_tier = eaf.tiers[gloss_tier_id]
                for gloss_ann_id, gloss_ann_contents in gloss_tier[0].items():
                    annotation_value = gloss_ann_contents[2]
                    if annotation_value in self.changes:
                        new_annotation_contents = (gloss_ann_contents[0], gloss_ann_contents[1],
                                                   self.changes[annotation_value][0],
                                                   gloss_ann_contents[3])
                        gloss_tier[0][gloss_ann_id] = new_annotation_contents

                        # Update or create refering meaning annotion
                        if gloss_ann_id in meaning_annotation_dict:
                            meaning_annotation_tuple = meaning_annotation_dict[gloss_ann_id]
                            gloss_ann_id = meaning_annotation_tuple[0]
                            meaning_tier[1][gloss_ann_id] = (gloss_ann_id,
                                                              meaning_annotation_tuple[1],
                                                              self.changes[annotation_value][1],
                                                              meaning_annotation_tuple[3],
                                                              )
                        else:
                            eaf.add_ref_annotation(meaning_tier_id, gloss_tier_id,
                                                   eaf.timeslots[gloss_ann_contents[0]],
                                                   self.changes[annotation_value][1])

if __name__ == "__main__":
    # -o Output directory; optional
    usage = "Usage: \n" + sys.argv[0] + \
            " -o <output directory>" + \
            " -e <excel with changes>"

    # Set default values
    output_dir = None
    excel_with_changes = None

    # Register command line arguments
    opt_list, file_list = getopt.getopt(sys.argv[1:], 'o:e:')
    for opt in opt_list:
        if opt[0] == '-o':
            output_dir = opt[1]
        if opt[0] == '-e':
            excel_with_changes = opt[1]

    # Check for errors and report
    errors = []
    if file_list is None or len(file_list) == 0:
        errors.append("No files or directories given.")
    if excel_with_changes is None:
        errors.append("No Excel with changes given.")

    if len(errors) != 0:
        print("Errors:")
        print("\n".join(errors))
        print(usage)
        exit(1)

    # Report registered options
    print("OPTIONS", file=sys.stderr)
    print("Files: " + ", ".join(file_list), file=sys.stderr)
    if output_dir is not None:
        print("Output directory: " + output_dir, file=sys.stderr)
    print("Excel with changes: " + excel_with_changes, file=sys.stderr)

    # Build and run
    file_collection_processor = FileCollectionProcessor(file_list, output_dir=output_dir, extensions_to_process=["eaf"])
    glossChanger = GlossChanger(excel_with_changes)
    file_collection_processor.add_file_processor(glossChanger)
    file_collection_processor.run()


