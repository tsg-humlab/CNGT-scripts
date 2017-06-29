#!/usr/bin/python

"""
Script to extract metadata from an Excel file.
"""

import sys, json
from openpyxl import load_workbook

wb = load_workbook(sys.argv[1])

# Participants
sheet_participants = wb['Participant metadata']

participants = {}
header = [cell.value for cell in sheet_participants['A1':'E1'][0]]
for row in sheet_participants.iter_rows(min_row=2, max_col=5, max_row=10):
    row_as_list = [cell.value for cell in row]

    participants[row_as_list[0]] = dict(zip(header[1:], row_as_list[1:]))


# Sessions
sheet_sessions = wb['Session metadata']

sessions = {}
header = [cell.value for cell in sheet_sessions['A1':'E1'][0]]
for row in sheet_sessions.iter_rows(min_row=2, max_col=5, max_row=10):
    row_as_list = [cell.value for cell in row]

    sessions[row_as_list[1]] = dict(zip(header[0:1]+header[2:], row_as_list[0:1]+row_as_list[2:]))

metadata = {
    'participants': participants,
    'sessions': sessions
}

# Output
print(json.dumps(metadata, sort_keys=True, indent=4))