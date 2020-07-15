#!/usr/bin/python

"""
Script to search and replace strings in CNGT EAFs.
"""

import sys
import getopt
from openpyxl import load_workbook
from CNGT_scripts.python.filecollectionprocessing.filecollectionprocessor import FileCollectionProcessor
from CNGT_scripts.python.filecollectionprocessing.eafprocessor import EafProcessor

class SearchReplace(EafProcessor):
    """
    
    """

    def __init__(self, excel_with_changes, first_row=1, search_column='A', replace_column='B'):
        # Reading the Excel with changes
        self.changes = dict()
        self.read_excel_with_changes(excel_with_changes, first_row, search_column, replace_column, self.changes)

    @staticmethod
    def read_excel_with_changes(excel_with_changes, first_row, search_column, replace_column, changes):
        """
        Reads the Excel containing changes and puts them in a dictionary
        :param excel_with_changes: 
        :param first_row: 
        :param search_column: 
        :param replace_column:
        :param changes:
        :return: 
        """
        # Load the workbook
        workbook = load_workbook(filename=excel_with_changes)
        for sheet_name in workbook.get_sheet_names():
            print("Sheet: ", sheet_name)
            search_replace_values = []
            worksheet = workbook[sheet_name]
            for row_index in range(first_row, worksheet.max_row + 1):
                search_value = worksheet[search_column + str(row_index)].value
                replace_value = worksheet[replace_column + str(row_index)].value
                search_replace_values.append((search_value, replace_value))
            changes[sheet_name] = sorted(search_replace_values,
                                         key=lambda search_replace_tuple: search_replace_tuple[0],
                                         reverse=True)
        print("Sheets: ", ", ".join(changes.keys()), str(workbook.get_sheet_names()))

    def process_eaf(self, eaf, file_name):
        print("EAF file: " + file_name)
        for lingtype in self.changes:
            print("Linguistic type: ", lingtype)
            for tier_name in eaf.get_tier_ids_for_linguistic_type(lingtype):
                print("Tier: ", tier_name)
                tier = eaf.tiers[tier_name]
                for ann_id, ann_contents in tier[0].items():
                    new_value = ann_contents[2]
                    for search_replace_tuple in self.changes[lingtype]:
                        new_value = new_value.replace(search_replace_tuple[0], search_replace_tuple[1])
                    tier[0][ann_id] = (ann_contents[0], ann_contents[1], new_value, ann_contents[3], ann_contents[4])

if __name__ == "__main__":
    # -o Output directory; optional
    usage = "Usage: \n" + sys.argv[0] + \
            " -o <output directory>" + \
            " -e <excel with changes>" + \
            " -r <first data row>"

    # Set default values
    output_dir = None
    excel_with_changes = None
    first_data_row = None

    # Register command line arguments
    opt_list, file_list = getopt.getopt(sys.argv[1:], 'o:e:r:')
    for opt in opt_list:
        if opt[0] == '-o':
            output_dir = opt[1]
        if opt[0] == '-e':
            excel_with_changes = opt[1]
        if opt[0] == '-r':
            first_data_row = int(opt[1])

    # Check for errors and report
    errors = []
    if file_list is None or len(file_list) == 0:
        errors.append("No files or directories given.")
    if excel_with_changes is None:
        errors.append("No Excel with changes given.")
    if first_data_row is None:
        first_data_row = 1

    if len(errors) != 0:
        print("Errors:")
        print("\n".join(errors))
        print(usage)
        exit(1)

    # Report registered options
    print("OPTIONS", file=sys.stderr)
    print("Files: " + ", ".join(file_list), file=sys.stderr)
    if output_dir is not None:
        print("Output directory: " + output_dir, file=sys.stderr)
    print("Excel with changes: " + excel_with_changes, file=sys.stderr)

    # Build and run
    file_collection_processor = FileCollectionProcessor(file_list, output_dir=output_dir, extensions_to_process=["eaf"])
    glossChanger = SearchReplace(excel_with_changes, first_row=first_data_row)
    file_collection_processor.add_file_processor(glossChanger)
    file_collection_processor.run()

