#!/usr/bin/python

"""
Script to add Signbank as a lexicon to the CNGT EAFs.
"""

import sys
import getopt
from openpyxl import load_workbook
from CNGT_scripts.python.filecollectionprocessing.filecollectionprocessor import FileCollectionProcessor
from CNGT_scripts.python.filecollectionprocessing.eafprocessor import EafProcessor

# Settings


class GlossChanger(EafProcessor):
    """
    
    """

    def __init__(self, excel_with_changes, first_row = 1, old_gloss_column = 'A', new_gloss_column = 'B',
                 meaning_column = 'C', ecv_url='https://signbank.science.ru.nl/static/ecv/ngt.ecv'):
        # Reading the ECV
        self.gloss_ids = self.read_ecv(ecv_url)
        # Reading the Excel with changes
        self.changes = dict()
        self.read_excel_with_changes(excel_with_changes, first_row, old_gloss_column, new_gloss_column, meaning_column,
                                     self.changes)

    def read_ecv(self, ecv_url):
        from lxml import etree
        import requests
        ecv_str = requests.get(ecv_url)
        ecv = etree.fromstring(ecv_str.content)

        gloss_ids = {}
        for entry in ecv.xpath('//CV_ENTRY_ML'):
            gloss = entry.xpath('./CVE_VALUE[contains(@LANG_REF, "nld")]/text()')[0]
            id = entry.get('CVE_ID')
            gloss_ids[gloss] = id

        return gloss_ids

    @staticmethod
    def read_excel_with_changes(excel_with_changes, first_row, old_gloss_column, new_gloss_column,
                                meaning_column, changes):
        """
        Reads the Excel containing changes and puts them in a dictionary
        :param excel_with_changes: 
        :param first_row: 
        :param old_gloss_column: 
        :param new_gloss_column: 
        :param meaning_column: 
        :return: 
        """
        # Load the workbook
        workbook = load_workbook(filename=excel_with_changes)
        # Take the first worksheet
        worksheet = workbook[workbook.get_sheet_names()[0]]
        for row_index in range(first_row, worksheet.max_row+1):
            try:
                old_gloss = worksheet[old_gloss_column+str(row_index)].value
                new_gloss = worksheet[new_gloss_column+str(row_index)].value
                meaning = worksheet[meaning_column+str(row_index)].value
                if (old_gloss is not None and old_gloss is not "" and
                    new_gloss is not None and new_gloss is not "" and
                    meaning is not None and meaning is not "" and
                    old_gloss not in changes):
                    changes[old_gloss] = (new_gloss, meaning)
            except Exception:
                print("Unable to process row " + str(row_index) + ". Error message: " + sys.exc_info())

    def process_eaf(self, eaf, file_name):
        print("EAF file: " + file_name)
        for subject_id in [1, 2]:
            for hand in ['L', 'R']:
                # Handle Meaning tier
                meaning_tier_id = 'Meaning' + hand + ' S' + str(subject_id)
                meaning_tier = eaf.tiers[meaning_tier_id]
                meaning_annotation_dict = {}
                for meaning_ann_id, meaning_ann_contents in meaning_tier[1].items():
                    # Switch annotation id and refering id. This is possible because
                    # the stereo type of Meaning tiers is symbolic association.
                    meaning_annotation_dict[meaning_ann_contents[0]] = (meaning_ann_id,
                                                                       meaning_ann_contents[1],
                                                                       meaning_ann_contents[2],
                                                                       meaning_ann_contents[3],
                                                                       )

                # Handle Gloss tier
                gloss_tier_id = 'Gloss' + hand + ' S' + str(subject_id)
                gloss_tier = eaf.tiers[gloss_tier_id]
                for gloss_ann_id, gloss_ann_contents in gloss_tier[0].items():
                    annotation_value = gloss_ann_contents[2]
                    if annotation_value in self.changes:
                        new_value = self.changes[annotation_value][0]
                        new_cve_ref = self.gloss_ids.get(new_value, None)
                        new_annotation_contents = (gloss_ann_contents[0], gloss_ann_contents[1],
                                                   new_value,
                                                   gloss_ann_contents[3],
                                                   new_cve_ref)
                        gloss_tier[0][gloss_ann_id] = new_annotation_contents

                        # Update or create refering meaning annotion
                        new_meaning = self.changes[annotation_value][1]
                        # print(gloss_ann_id)
                        if gloss_ann_id in meaning_annotation_dict:
                            # print(gloss_ann_id)
                            meaning_annotation_tuple = meaning_annotation_dict[gloss_ann_id]
                            meaning_ann_id = meaning_annotation_tuple[0]
                            meaning_tier[1][meaning_ann_id] = (gloss_ann_id,
                                                             new_meaning,
                                                             meaning_annotation_tuple[2],
                                                             meaning_annotation_tuple[3])
                        else:
                            eaf.add_ref_annotation(meaning_tier_id, gloss_tier_id,
                                                   eaf.timeslots[gloss_ann_contents[0]],
                                                   new_meaning)

if __name__ == "__main__":
    # -o Output directory; optional
    usage = "Usage: \n" + sys.argv[0] + \
            " -o <output directory>" + \
            " -e <excel with changes>" + \
            " -r <first data row>"

    # Set default values
    output_dir = None
    excel_with_changes = None

    # Register command line arguments
    opt_list, file_list = getopt.getopt(sys.argv[1:], 'o:e:r:')
    for opt in opt_list:
        if opt[0] == '-o':
            output_dir = opt[1]
        if opt[0] == '-e':
            excel_with_changes = opt[1]
        if opt[0] == '-r':
            first_data_row = int(opt[1])

    # Check for errors and report
    errors = []
    if file_list is None or len(file_list) == 0:
        errors.append("No files or directories given.")
    if excel_with_changes is None:
        errors.append("No Excel with changes given.")
    if first_data_row is None:
        first_data_row = 1

    if len(errors) != 0:
        print("Errors:")
        print("\n".join(errors))
        print(usage)
        exit(1)

    # Report registered options
    print("OPTIONS", file=sys.stderr)
    print("Files: " + ", ".join(file_list), file=sys.stderr)
    if output_dir is not None:
        print("Output directory: " + output_dir, file=sys.stderr)
    print("Excel with changes: " + excel_with_changes, file=sys.stderr)

    # Build and run
    file_collection_processor = FileCollectionProcessor(file_list, output_dir=output_dir, extensions_to_process=["eaf"])
    glossChanger = GlossChanger(excel_with_changes, first_row=first_data_row)
    file_collection_processor.add_file_processor(glossChanger)
    file_collection_processor.run()


